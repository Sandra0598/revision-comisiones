"""Generacion del Excel final de revision con openpyxl.

Crea las pestanas: Configuracion, Resumen general, una por comercial, Eva, Sara,
Concurso Pareja Eva-Sara e Incidencias. Aplica estilos basicos: cabeceras en
negrita, autofiltro, ancho automatico y formatos de moneda/porcentaje.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import ComercialResult, Incidencia, ParejaResult
from app.utils.normalize import names_match, normalize_text

# Estilos compartidos.
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
LABEL_FONT = Font(bold=True)
# Estilos para destacar el TOTAL y que se localice de un vistazo.
TOTAL_FONT = Font(bold=True, size=16, color="1F4E78")
TOTAL_LABEL_FONT = Font(bold=True, size=13, color="1F4E78")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_HEADER_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
GRAVEDAD_FILL = {
    "alta": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "media": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    "baja": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}

MONEY_FMT = '#,##0.00 "€"'
PCT_FMT = '0.00"%"'

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre",
    12: "Diciembre",
}


def _style_header_row(ws: Worksheet, row_idx: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autofit(ws: Worksheet, max_width: int = 45) -> None:
    """Ajusta el ancho de columnas segun el contenido."""
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column
            if length > widths.get(col, 0):
                widths[col] = length
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), max_width)


def _write_table(
    ws: Worksheet,
    headers: List[str],
    rows: List[List[Any]],
    start_row: int = 1,
    money_cols: Optional[List[int]] = None,
    pct_cols: Optional[List[int]] = None,
) -> int:
    """Escribe una tabla con cabecera estilizada y autofiltro. Devuelve fila final."""
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    _style_header_row(ws, start_row, len(headers))

    r = start_row
    for data in rows:
        r += 1
        for c, value in enumerate(data, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = THIN_BORDER
            if c in money_cols and isinstance(value, (int, float)):
                cell.number_format = MONEY_FMT
            elif c in pct_cols and isinstance(value, (int, float)):
                cell.number_format = PCT_FMT

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{start_row}:{last_col}{max(r, start_row)}"
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return r


# ---------------------------------------------------------------------------
# Pestana Configuracion
# ---------------------------------------------------------------------------

def _write_config_sheet(ws: Worksheet, config: Dict[str, Any]) -> None:
    ws.cell(row=1, column=1, value="Configuracion utilizada").font = TITLE_FONT
    headers = ["Parametro", "Valor"]
    rows: List[List[Any]] = []
    for key, value in config.items():
        if isinstance(value, (list, dict)):
            value = str(value)
        rows.append([key, value])
    _write_table(ws, headers, rows, start_row=3)
    _autofit(ws)


# ---------------------------------------------------------------------------
# Pestana Resumen general
# ---------------------------------------------------------------------------

# "Total esperado" va en 2.ª posicion (justo tras el nombre) para localizarlo
# de un vistazo y se resalta con color propio.
RESUMEN_HEADERS = [
    "Comercial", "Total esperado", "Ventas computables totales", "Ventas Pizarra 1",
    "Ventas Pizarra 2 reales", "ADW computables", "Puntos ADW Pizarra 2",
    "RNS computables", "Puntos RNS Pizarra 2", "Pizarra 2 ajustada",
    "% concurso", "Concurso €", "Ventas nuevas 499+", "Premio ventas nuevas 499+",
    "Ventas para rappel bonus", "Rappel bonus general",
    "Ventas para rappel individual", "Nivel rappel individual",
    "Mes consecutivo individual", "Rappel individual", "Super concurso pareja",
    "Comision declarada", "Diferencia", "Estado", "Alertas",
]


def _resumen_row(res: ComercialResult) -> List[Any]:
    return [
        res.comercial,
        res.total_final,
        res.ventas_computables_totales,
        res.pizarra_1_final,
        res.pizarra_2_real,
        res.adw_computables,
        res.puntos_adw,
        res.rns_computables,
        res.puntos_rns,
        res.pizarra_2_ajustada,
        res.porcentaje_concurso,
        res.importe_concurso_obtenido,
        res.ventas_nuevas_499,
        res.premio_ventas_nuevas_499,
        res.ventas_para_rappel_bonus,
        res.rappel_bonus_general,
        res.ventas_para_rappel_individual,
        res.nivel_rappel_individual or "",
        res.mes_consecutivo_individual if res.mes_consecutivo_individual is not None else "",
        res.rappel_individual,
        res.super_concurso_pareja,
        res.comision_declarada if res.comision_declarada is not None else "",
        res.diferencia if res.diferencia is not None else "",
        res.estado,
        "; ".join(res.alertas),
    ]


def _write_resumen_sheet(ws: Worksheet, resultados: List[ComercialResult]) -> None:
    ws.cell(row=1, column=1, value="Resumen general de comisiones").font = TITLE_FONT
    rows = [_resumen_row(r) for r in resultados]
    # Columnas (1-indexadas) con formato moneda y porcentaje.
    total_col = RESUMEN_HEADERS.index("Total esperado") + 1  # = 2
    money_cols = [total_col, 12, 14, 16, 20, 21, 22, 23]
    pct_cols = [11]
    last = _write_table(ws, RESUMEN_HEADERS, rows, start_row=3, money_cols=money_cols, pct_cols=pct_cols)

    # Resaltar la columna del Total: cabecera naranja y celdas en negrita/oro.
    hdr = ws.cell(row=3, column=total_col)
    hdr.fill = TOTAL_HEADER_FILL
    for r in range(4, last + 1):
        cell = ws.cell(row=r, column=total_col)
        cell.font = LABEL_FONT
        cell.fill = TOTAL_FILL

    # Resaltar estados.
    estado_col = RESUMEN_HEADERS.index("Estado") + 1
    for r in range(4, last + 1):
        cell = ws.cell(row=r, column=estado_col)
        if cell.value == "Declarada de menos":
            cell.fill = GRAVEDAD_FILL["alta"]
        elif cell.value == "Declarada de mas":
            cell.fill = GRAVEDAD_FILL["media"]
        elif cell.value == "Correcta":
            cell.fill = GRAVEDAD_FILL["baja"]
    _autofit(ws)


# ---------------------------------------------------------------------------
# Pestana individual por comercial
# ---------------------------------------------------------------------------

DETALLE_HEADERS = [
    "Fecha", "Comercial", "Cliente", "Categoria venta (orig.)", "Categoria normalizada",
    "Servicio contratado",
    "Pizarra original", "Pizarra normalizada", "Importe", "Es venta 859",
    "Valor computable", "Computa", "Es Pizarra 1", "Es Pizarra 2",
    "Es cartera (ADW)", "Numero acumulado ADW", "Punto ADW en esta fila",
    "Es subida precio (RNS)", "Numero acumulado RNS", "Punto RNS en esta fila",
    "Es igual de precio", "Es bajada de precio", "Es upselling",
    "Es venta fria", "Es venta nueva 499+", "Comision declarada",
    "Observaciones", "Alertas fila",
]


def _fmt_fecha(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return value if value is not None else ""


def _write_resumen_bloque(ws: Worksheet, res: ComercialResult, start_row: int) -> int:
    """Escribe el bloque-resumen de magnitudes de la comercial.

    Arriba del todo pone un banner grande con el TOTAL para localizarlo de un
    vistazo; debajo, el resto de magnitudes.
    """
    ws.cell(row=start_row, column=1, value=f"Resumen {res.comercial}").font = TITLE_FONT

    # Banner destacado del TOTAL, justo bajo el titulo.
    banner_row = start_row + 1
    lc = ws.cell(row=banner_row, column=1, value="TOTAL")
    lc.font = TOTAL_LABEL_FONT
    lc.fill = TOTAL_FILL
    lc.alignment = Alignment(vertical="center")
    lc.border = THIN_BORDER
    vc = ws.cell(row=banner_row, column=2, value=res.total_final)
    vc.font = TOTAL_FONT
    vc.fill = TOTAL_FILL
    vc.alignment = Alignment(vertical="center")
    vc.border = THIN_BORDER
    if isinstance(res.total_final, (int, float)):
        vc.number_format = MONEY_FMT
    ws.row_dimensions[banner_row].height = 24

    pares = [
        ("Ventas computables totales", res.ventas_computables_totales, False),
        ("Pizarra 1 final", res.pizarra_1_final, False),
        ("Pizarra 2 real", res.pizarra_2_real, False),
        ("Puntos ADW", res.puntos_adw, False),
        ("Puntos RNS", res.puntos_rns, False),
        ("Pizarra 2 ajustada", res.pizarra_2_ajustada, False),
        ("Porcentaje concurso", res.porcentaje_concurso, "pct"),
        ("Importe concurso obtenido", res.importe_concurso_obtenido, True),
        ("Premio ventas nuevas 499+", res.premio_ventas_nuevas_499, True),
        ("Rappel bonus general", res.rappel_bonus_general, True),
        ("Rappel individual", res.rappel_individual, True),
        ("Super concurso pareja", res.super_concurso_pareja, True),
        ("Total final", res.total_final, True),
        ("Diferencia", res.diferencia if res.diferencia is not None else "", True),
        ("Ventas frias computables", res.ventas_frias_computables, False),
        ("Upselling computable", res.upselling_computable, False),
        ("Upselling equivalente fria", res.upselling_equivalente_fria, False),
        ("Ventas equivalentes totales", res.ventas_equivalentes_totales, False),
    ]
    r = banner_row
    for label, value, fmt in pares:
        r += 1
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        vc = ws.cell(row=r, column=2, value=value)
        if fmt == "pct" and isinstance(value, (int, float)):
            vc.number_format = PCT_FMT
        elif fmt is True and isinstance(value, (int, float)):
            vc.number_format = MONEY_FMT
    return r


def _write_comercial_sheet(ws: Worksheet, res: ComercialResult) -> None:
    # Bloque resumen arriba.
    last_resumen = _write_resumen_bloque(ws, res, start_row=1)
    start_tabla = last_resumen + 2
    ws.cell(row=start_tabla, column=1, value="Detalle de ventas").font = TITLE_FONT

    rows: List[List[Any]] = []
    for f in res.filas:
        rows.append([
            _fmt_fecha(f.get("Fecha")),
            f.get("Comercial", ""),
            f.get("Cliente", ""),
            f.get("Tipo venta original", ""),
            f.get("Tipo venta normalizado", ""),
            f.get("Servicio contratado", ""),
            f.get("Pizarra original", ""),
            f.get("Pizarra normalizada") if f.get("Pizarra normalizada") is not None else "",
            f.get("Importe") if f.get("Importe") is not None else "",
            "Si" if f.get("Es venta 859") else "No",
            f.get("Valor computable", ""),
            "Si" if f.get("Computa") else "No",
            "Si" if f.get("Es Pizarra 1") else "No",
            "Si" if f.get("Es Pizarra 2") else "No",
            "Si" if f.get("Es cartera (ADW)") else "No",
            f.get("Numero acumulado ADW", ""),
            f.get("Punto ADW en esta fila", 0),
            "Si" if f.get("Es subida precio RNS") else "No",
            f.get("Numero acumulado RNS", ""),
            f.get("Punto RNS en esta fila", 0),
            "Si" if f.get("Es igual de precio") else "No",
            "Si" if f.get("Es bajada de precio") else "No",
            "Si" if f.get("Es upselling") else "No",
            "Si" if f.get("Es venta fria") else "No",
            "Si" if f.get("Es venta nueva 499+") else "No",
            f.get("Comision declarada") if f.get("Comision declarada") is not None else "",
            f.get("Observaciones", ""),
            f.get("Alertas fila", ""),
        ])
    # Importe col=8, comision declarada col=25.
    _write_table(ws, DETALLE_HEADERS, rows, start_row=start_tabla + 1, money_cols=[8, 25])
    # No congelar (hay bloque arriba): ajustar freeze a la cabecera de tabla.
    ws.freeze_panes = ws.cell(row=start_tabla + 2, column=1)
    _autofit(ws)


# ---------------------------------------------------------------------------
# Pestana Concurso Pareja Eva-Sara
# ---------------------------------------------------------------------------

def _write_pareja_sheet(ws: Worksheet, pareja: Optional[ParejaResult], config: Dict[str, Any]) -> None:
    pareja_nombres = config.get("pareja_fija", ["Eva", "Sara"])
    n1 = pareja_nombres[0] if pareja_nombres else "Eva"
    n2 = pareja_nombres[1] if len(pareja_nombres) > 1 else "Sara"
    ws.cell(row=1, column=1, value=f"Concurso Pareja {n1}-{n2}").font = TITLE_FONT

    if pareja is None:
        ws.cell(row=3, column=1, value="No se pudo calcular el concurso de pareja (datos incompletos).")
        _autofit(ws)
        return

    pares = [
        (f"Ventas computables {n1}", pareja.ventas_computables_eva, False),
        (f"Ventas computables {n2}", pareja.ventas_computables_sara, False),
        ("Ventas brutas pareja", pareja.ventas_brutas_pareja, False),
        ("Ventas obligatorias restadas", pareja.ventas_obligatorias_restadas, False),
        ("Ventas validas pareja", pareja.ventas_validas_pareja, False),
        (f"Porcentaje concurso {n1}", pareja.porcentaje_concurso_eva, "pct"),
        (f"Porcentaje concurso {n2}", pareja.porcentaje_concurso_sara, "pct"),
        ("Aplica pareja", "Si" if pareja.aplica_pareja else "No", False),
        ("Nivel pareja actual", pareja.nivel_pareja_actual or "—", False),
        ("Mes consecutivo pareja", pareja.mes_consecutivo_pareja if pareja.mes_consecutivo_pareja is not None else "No informado", False),
        ("Premio por persona", pareja.premio_por_persona, True),
        ("Autorizacion Raul", pareja.autorizacion_raul, False),
        ("Alerta", pareja.alerta or "", False),
        (f"Importe final {n1}", pareja.importe_final_eva, True),
        (f"Importe final {n2}", pareja.importe_final_sara, True),
    ]
    r = 2
    for label, value, fmt in pares:
        r += 1
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        vc = ws.cell(row=r, column=2, value=value)
        if fmt == "pct" and isinstance(value, (int, float)):
            vc.number_format = PCT_FMT
        elif fmt is True and isinstance(value, (int, float)):
            vc.number_format = MONEY_FMT
    _autofit(ws)


# ---------------------------------------------------------------------------
# Pestana Incidencias
# ---------------------------------------------------------------------------

def _write_incidencias_sheet(ws: Worksheet, incidencias: List[Incidencia]) -> None:
    ws.cell(row=1, column=1, value="Incidencias y alertas").font = TITLE_FONT
    headers = ["Fila original", "Comercial", "Tipo incidencia", "Descripcion", "Gravedad"]
    rows = [
        [
            inc.fila_original if inc.fila_original is not None else "",
            inc.comercial or "",
            inc.tipo,
            inc.descripcion,
            inc.gravedad,
        ]
        for inc in incidencias
    ]
    last = _write_table(ws, headers, rows, start_row=3)
    # Colorear por gravedad.
    for i, inc in enumerate(incidencias):
        row_idx = 4 + i
        fill = GRAVEDAD_FILL.get(inc.gravedad)
        if fill:
            ws.cell(row=row_idx, column=5).fill = fill
    _autofit(ws)


# ---------------------------------------------------------------------------
# Saneado de nombres de hoja
# ---------------------------------------------------------------------------

_INVALID_SHEET_CHARS = set(r"[]:*?/\\")


def _safe_sheet_name(name: str, used: set) -> str:
    clean = "".join(c for c in str(name) if c not in _INVALID_SHEET_CHARS).strip()
    clean = clean[:31] or "Hoja"
    base = clean
    i = 2
    while clean.lower() in used:
        suffix = f" {i}"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


# ---------------------------------------------------------------------------
# Entrada principal
# ---------------------------------------------------------------------------

def build_workbook(
    resultados: Dict[str, ComercialResult],
    pareja: Optional[ParejaResult],
    incidencias: List[Incidencia],
    config: Dict[str, Any],
    output_path: str,
) -> str:
    """Construye y guarda el Excel final. Devuelve la ruta del archivo."""
    wb = Workbook()
    used_names: set = set()

    # 1. Configuracion (primera hoja).
    ws_config = wb.active
    ws_config.title = _safe_sheet_name("Configuracion", used_names)
    _write_config_sheet(ws_config, config)

    # Ordenar comerciales: primero el resto, luego Eva y Sara al final.
    pareja_nombres = config.get("pareja_fija", ["Eva", "Sara"])
    lista = list(resultados.values())

    def es_pareja(res: ComercialResult) -> bool:
        return any(names_match(res.comercial, p) for p in pareja_nombres)

    resto = sorted([r for r in lista if not es_pareja(r)], key=lambda r: normalize_text(r.comercial))
    pareja_res = [r for r in lista if es_pareja(r)]
    # Ordenar pareja segun el orden de pareja_fija.
    pareja_res.sort(
        key=lambda r: next(
            (i for i, p in enumerate(pareja_nombres) if names_match(r.comercial, p)), 99
        )
    )

    # 2. Resumen general.
    ws_resumen = wb.create_sheet(_safe_sheet_name("Resumen general", used_names))
    _write_resumen_sheet(ws_resumen, resto + pareja_res)

    # 3. Una pestana por comercial (resto), 4-5. Eva y Sara.
    for res in resto + pareja_res:
        ws = wb.create_sheet(_safe_sheet_name(res.comercial, used_names))
        _write_comercial_sheet(ws, res)

    # 6. Concurso Pareja.
    n1 = pareja_nombres[0] if pareja_nombres else "Eva"
    n2 = pareja_nombres[1] if len(pareja_nombres) > 1 else "Sara"
    ws_pareja = wb.create_sheet(_safe_sheet_name(f"Concurso Pareja {n1}-{n2}", used_names))
    _write_pareja_sheet(ws_pareja, pareja, config)

    # 7. Incidencias.
    ws_inc = wb.create_sheet(_safe_sheet_name("Incidencias", used_names))
    _write_incidencias_sheet(ws_inc, incidencias)

    wb.save(output_path)
    return output_path


def build_output_filename(mes: int, anio: int) -> str:
    """Construye el nombre del archivo de salida."""
    mes_nombre = MESES_ES.get(mes, str(mes))
    return f"Revision_Comisiones_{mes_nombre}_{anio}.xlsx"
