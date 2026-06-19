"""Test end-to-end del modelo de comisiones con los filtros de negocio.

Reglas verificadas:
  - Solo se revisan las 8 comerciales validas.
  - Solo se revisan filas con pizarra 1/2, subida/igual/bajada de precio o
    categoria 'cartera'.
  - 'cartera' -> ADW (1 punto Pizarra 2 cada 3).
  - 'subida precio' -> RNS (1 punto Pizarra 2 cada 4).
  - 'igual de precio' / 'bajada de precio' -> se muestran pero no computan.
  - Eva y Sara: el rappel en pareja sustituye al rappel individual.

Ejecutar: python test_flow.py
"""

import io
import os

import pandas as pd

from app.data.config_defaults import merge_config
from app.data.tables import (
    determinar_nivel_pareja,
    lookup_rappel_bonus_general,
    lookup_rappel_individual,
    lookup_super_concurso_pareja,
)
from app.services.commission_engine import analizar
from app.services.excel_reader import build_rows, infer_mes_anio, read_excel
from app.services.excel_writer import build_output_filename, build_workbook
from app.services.validators import (
    adw_punto_fila,
    fila_incluida,
    rns_punto_fila,
    valor_computable,
)
from app.utils.normalize import (
    detect_movimiento_precio,
    detect_pizarra,
    is_cartera,
    normalize_amount,
    normalize_text,
)


def test_unidades():
    print("== Tests unitarios ==")
    assert normalize_text("  Pizarra  Uno ") == "pizarra uno"
    assert normalize_amount("859 €") == 859.0
    assert normalize_amount("1.234,56") == 1234.56
    assert detect_pizarra("1") == 1
    assert detect_pizarra("2") == 2
    assert detect_movimiento_precio("Subida precio") == "subida"
    assert detect_movimiento_precio("Igual de precio") == "igual"
    assert detect_movimiento_precio("Bajada precio") == "bajada"
    assert detect_movimiento_precio("1") is None
    assert is_cartera("Cartera")
    assert not is_cartera("Pizarra 1")

    # Filtro de fila incluida.
    assert fila_incluida({"pizarra": "1", "tipo_venta": "x"})
    assert fila_incluida({"pizarra": "Subida precio", "tipo_venta": "x"})
    assert fila_incluida({"pizarra": "", "tipo_venta": "Cartera"})
    assert not fila_incluida({"pizarra": "", "tipo_venta": "otros"})

    cfg = merge_config({})
    assert valor_computable(859.0, cfg) == 2
    assert [adw_punto_fila(i) for i in range(1, 10)] == [0, 0, 1, 0, 0, 1, 0, 0, 1]
    assert [rns_punto_fila(i) for i in range(1, 13)] == [0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 1]

    assert lookup_rappel_bonus_general(8) == 40
    assert lookup_rappel_bonus_general(58) == 331
    nivel, importe = lookup_rappel_individual(31, "1")
    assert nivel == "senior" and importe == 425
    assert determinar_nivel_pareja(47) == "Baby"
    assert lookup_super_concurso_pareja("Baby", "1") == 75
    print("OK tests unitarios")


def crear_excel_demo() -> bytes:
    """Excel de ventas de ejemplo con las columnas pizarra y categoria venta."""
    filas = []

    # Eva (pareja). Columnas: Comercial, Fecha, Cliente, Categoria venta,
    # Pizarra, Importe, Comision declarada, Producto, Observaciones.
    for i in range(13):
        filas.append(["Eva", "05/04/2026", f"E{i}", "nueva", "1", 500, None, "Plan", ""])
    for i in range(12):
        filas.append(["Eva", "06/04/2026", f"E2{i}", "nueva", "2", 859, None, "Plan", ""])
    for i in range(3):
        filas.append(["Eva", "07/04/2026", f"EC{i}", "Cartera", "", 300, None, "ADW", ""])
    for i in range(2):
        filas.append(["Eva", "08/04/2026", f"EI{i}", "renovacion", "Igual de precio", 400, None, "Plan", ""])

    # Sara (pareja).
    for i in range(13):
        filas.append(["Sara", "05/04/2026", f"S{i}", "nueva", "1", 550, None, "Plan", ""])
    for i in range(12):
        filas.append(["Sara", "06/04/2026", f"S2{i}", "nueva", "2", 859, None, "Plan", ""])
    for i in range(4):
        filas.append(["Sara", "07/04/2026", f"SR{i}", "renovacion", "Subida precio", 200, None, "RNS", ""])
    for i in range(20):
        filas.append(["Sara", "09/04/2026", f"S3{i}", "nueva", "1", 500, None, "Plan", ""])

    # Laura (valida pero no llega al 100%) con comision declarada.
    for i in range(5):
        filas.append(["Laura", "10/04/2026", f"L{i}", "nueva", "1", 400, 100, "Plan", ""])

    # Pedro: comercial NO valida -> debe filtrarse por completo.
    for i in range(3):
        filas.append(["Pedro", "10/04/2026", f"P{i}", "nueva", "1", 500, None, "Plan", ""])

    # Fila fuera de filtro (pizarra vacia y categoria no 'cartera').
    filas.append(["Eva", "11/04/2026", "EX", "otros", "", 100, None, "Plan", ""])

    cols = ["Comercial", "Fecha", "Cliente", "Categoria venta", "Pizarra",
            "Importe", "Comisión declarada", "Producto", "Observaciones"]
    df = pd.DataFrame(filas, columns=cols)
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    return buf.getvalue()


def test_flujo():
    print("== Test flujo completo ==")
    content = crear_excel_demo()
    df, mapping, inc_lec = read_excel(content)
    print("Mapeo:", {k: v for k, v in mapping.items() if v})
    assert mapping["tipo_venta"] == "Categoria venta", mapping["tipo_venta"]
    assert mapping["pizarra"] == "Pizarra"

    rows, inc_filas = build_rows(df, mapping)
    mes, anio = infer_mes_anio(rows)
    assert mes == 4 and anio == 2026

    cfg = merge_config({
        "importe_concurso_mensual": 1000,
        "mes_consecutivo_pareja": 1,
    })

    resultados, pareja, inc_calc = analizar(rows, cfg)
    incidencias = inc_lec + inc_filas + inc_calc

    # Pedro no debe existir.
    assert "pedro" not in resultados, "Pedro no deberia revisarse"

    eva = resultados["eva"]
    sara = resultados["sara"]
    laura = resultados["laura"]

    print(f"Eva: comp={eva.ventas_computables_totales} P1={eva.pizarra_1_final} "
          f"P2real={eva.pizarra_2_real} ADW={eva.adw_computables} ptsADW={eva.puntos_adw} "
          f"P2aj={eva.pizarra_2_ajustada} %={eva.porcentaje_concurso} "
          f"499={eva.ventas_nuevas_499} rapBonus={eva.rappel_bonus_general} "
          f"rapInd={eva.rappel_individual} pareja={eva.super_concurso_pareja} total={eva.total_final}")
    print(f"Sara: comp={sara.ventas_computables_totales} RNS={sara.rns_computables} "
          f"ptsRNS={sara.puntos_rns} P2aj={sara.pizarra_2_ajustada} %={sara.porcentaje_concurso} "
          f"pareja={sara.super_concurso_pareja} total={sara.total_final}")
    print(f"Laura: comp={laura.ventas_computables_totales} %={laura.porcentaje_concurso} "
          f"premio499={laura.premio_ventas_nuevas_499} pareja={laura.super_concurso_pareja} "
          f"declarada={laura.comision_declarada} dif={laura.diferencia} estado={laura.estado}")

    # Eva: 13 P1 + 24 (12x859 doble) P2 + 3 cartera(ADW) = 40 computables.
    assert eva.ventas_computables_totales == 40, eva.ventas_computables_totales
    assert eva.pizarra_1_final == 13
    assert eva.pizarra_2_real == 24
    assert eva.adw_computables == 3 and eva.puntos_adw == 1
    assert eva.pizarra_2_ajustada == 25
    assert eva.porcentaje_concurso == 100
    assert eva.ventas_nuevas_499 == 25  # 13 P1>=499 + 12 P2>=499
    # Eva es pareja -> rappel individual sustituido por pareja.
    assert eva.rappel_individual == 0
    # Conserva rappel bonus general (15 ventas validas -> 100).
    assert eva.rappel_bonus_general == lookup_rappel_bonus_general(15)

    # Sara: subida precio -> RNS (4 -> 1 punto).
    assert sara.rns_computables == 4 and sara.puntos_rns == 1
    # 13 + 24 + 4 + 20 = 61 computables.
    assert sara.ventas_computables_totales == 61, sara.ventas_computables_totales

    # Pareja: 40 + 61 - 50 = 51 -> Baby -> 75/persona, ambas 100%.
    assert pareja is not None and pareja.aplica_pareja
    assert pareja.ventas_validas_pareja == 51
    assert pareja.nivel_pareja_actual == "Baby"
    assert pareja.premio_por_persona == 75
    assert eva.super_concurso_pareja == 75 and sara.super_concurso_pareja == 75

    # Laura: no llega al 100% -> sin premios; no es pareja.
    assert laura.porcentaje_concurso < 100
    assert laura.premio_ventas_nuevas_499 == 0
    assert laura.super_concurso_pareja == 0
    assert laura.estado == "Declarada de mas"

    # Incidencias de filtros.
    tipos = {i.tipo for i in incidencias}
    assert "comercial_fuera_de_lista" in tipos
    assert "filas_fuera_de_filtro" in tipos

    # Generar Excel.
    nombre = build_output_filename(mes, anio)
    out = os.path.join("tmp", "TEST__" + nombre)
    build_workbook(resultados, pareja, incidencias, cfg, out)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    print("Pestanas:", wb.sheetnames)
    assert "Eva" in wb.sheetnames and "Sara" in wb.sheetnames and "Laura" in wb.sheetnames
    assert "Pedro" not in wb.sheetnames

    print(f"Incidencias ({len(incidencias)}):")
    for inc in incidencias:
        print(f"  [{inc.gravedad}] {inc.tipo}: {inc.descripcion}")
    print("OK flujo completo")


if __name__ == "__main__":
    test_unidades()
    test_flujo()
    print("\nTODOS LOS TESTS PASARON")
